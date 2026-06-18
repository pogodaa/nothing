"""Импорт данных из xlsx (папка import/) в SQLite."""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import config
import database as db


def import_all(conn, import_dir: Path | None = None) -> None:
    """Загружает все xlsx из import/ в velodrive.db (первый запуск приложения)."""
    folder = import_dir or config.IMPORT_DIR
    if not folder.exists():
        raise FileNotFoundError(f"Папка import не найдена: {folder}")

    db.init_schema(conn)
    conn.execute("DELETE FROM order_items")
    conn.execute("DELETE FROM orders")
    conn.execute("DELETE FROM products")
    conn.execute("DELETE FROM users")
    conn.execute("DELETE FROM pickup_points")

    _import_pickup_points(conn, folder / "Пункты выдачи_import.xlsx")
    _import_users(conn, folder / "user_import.xlsx")
    _import_products(conn, folder / "Tovar.xlsx", folder)
    _import_orders(conn, folder / "Заказ_import.xlsx")
    conn.commit()


def _import_pickup_points(conn, path: Path) -> None:
    wb = load_workbook(path, read_only=False)
    rows = list(wb.active.iter_rows(values_only=True))
    point_id = 1
    for row in rows:
        address = row[0]
        if not address:
            continue
        conn.execute(
            "INSERT INTO pickup_points (id, address) VALUES (?, ?)",
            (point_id, str(address).strip()),
        )
        point_id += 1


def _import_users(conn, path: Path) -> None:
    wb = load_workbook(path, read_only=False)
    rows = list(wb.active.iter_rows(values_only=True))[1:]
    for row in rows:
        role, full_name, login, password = row[:4]
        if not login:
            continue
        conn.execute(
            """
            INSERT INTO users (role, full_name, login, password)
            VALUES (?, ?, ?, ?)
            """,
            (str(role).strip(), str(full_name).strip(), str(login).strip(), str(password)),
        )


def _import_products(conn, path: Path, import_dir: Path) -> None:
    wb = load_workbook(path, read_only=False)
    rows = list(wb.active.iter_rows(values_only=True))[1:]
    product_id = 1
    for row in rows:
        if not row[0]:
            continue
        photo_name = str(row[10] or "").strip()
        photo_path = _find_photo(import_dir, photo_name)
        conn.execute(
            """
            INSERT INTO products (
                id, article, name, unit, price, supplier, manufacturer,
                category, discount, quantity, description, photo_path
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                product_id,
                db.normalize_article(str(row[0])),
                str(row[1]).strip(),
                str(row[2]).strip(),
                float(row[3] or 0),
                str(row[4]).strip(),
                str(row[5]).strip(),
                str(row[6]).strip(),
                float(row[7] or 0),
                int(row[8] or 0),
                str(row[9] or "").strip(),
                photo_path,
            ),
        )
        product_id += 1


def _find_photo(import_dir: Path, photo_name: str) -> str:
    if not photo_name:
        return _photo_path_for_db(config.PICTURE_PLACEHOLDER)
    for variant in (photo_name, photo_name.upper(), photo_name.lower()):
        candidate = import_dir / variant
        if candidate.exists():
            return _photo_path_for_db(candidate)
        stem = Path(variant).stem
        for ext in (".JPG", ".jpg", ".png", ".PNG"):
            candidate = import_dir / f"{stem}{ext}"
            if candidate.exists():
                return _photo_path_for_db(candidate)
    return _photo_path_for_db(config.PICTURE_PLACEHOLDER)


def _photo_path_for_db(path: Path) -> str:
    """Путь относительно APP_DIR (import/... или images/...) — переносимо на Linux."""
    resolved = path.resolve()
    try:
        relative = resolved.relative_to(config.APP_DIR.resolve())
    except ValueError:
        return str(resolved)
    return relative.as_posix()


def _import_orders(conn, path: Path) -> None:
    wb = load_workbook(path, read_only=False)
    rows = list(wb.active.iter_rows(values_only=True))[1:]
    for row in rows:
        if row[0] is None:
            continue
        order_id = int(row[0])
        articles_text = db.normalize_article(str(row[1] or "").strip())
        order_date = db.format_date(row[2])
        delivery_date = db.format_date(row[3])
        pickup_point_id = int(row[4]) if row[4] is not None else None
        client_name = str(row[5] or "").strip()
        pickup_code = int(row[6]) if row[6] is not None else None
        status = str(row[7] or "").strip()
        conn.execute(
            """
            INSERT INTO orders (
                id, articles_text, order_date, delivery_date,
                pickup_point_id, client_name, pickup_code, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                order_id,
                articles_text,
                order_date,
                delivery_date,
                pickup_point_id,
                client_name,
                pickup_code,
                status,
            ),
        )
        for article, qty in db.parse_articles_text(articles_text):
            conn.execute(
                """
                INSERT INTO order_items (order_id, product_article, quantity)
                VALUES (?, ?, ?)
                """,
                (order_id, article, qty),
            )


def export_sql_script(db_path: Path, output_path: Path) -> None:
    """Создаёт database.sql из текущей БД (для сдачи модуля 1)."""
    import sqlite3
    import subprocess
    import sys

    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(db_path)
    import_all(conn)
    conn.close()

    with output_path.open("w", encoding="utf-8") as file:
        subprocess.run(
            [sys.executable, "-m", "sqlite3", str(db_path), ".dump"],
            stdout=file,
            check=False,
        )


if __name__ == "__main__":
    conn = db.get_connection()
    import_all(conn)
    conn.close()
    print("Импорт завершён:", config.DB_PATH)
